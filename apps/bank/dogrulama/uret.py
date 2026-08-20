# -*- coding: utf-8 -*-
"""
Kleaf Karbon Hesaplama Motoru — Doğrulama Çalışma Kitabı Üreteci
================================================================
Bu betik, src/lib/carbon/engine.ts içindeki TÜM hesaplamaları
bağımsız olarak teyit edebileceğiniz bir Excel (.xlsx) dosyası üretir.

Her sayfada:
  - Açıklama + motordaki fonksiyon adı
  - Test girdi verileri (engine.test.ts'teki bilinen değerler)
  - CANLI Excel formülü  (B sütunu — hücre referanslarından hesaplar)
  - Beklenen değer        (C sütunu — motorun ürettiği bilinen sonuç)
  - KONTROL               (D sütunu — otomatik "✓ EŞLEŞTİ" / "✗ FARK")
  - Formülün metni        (E sütunu — B'deki formülün okunabilir hâli)
  - Kaynak                (F sütunu — engine.ts fonksiyonu)

Canlı formül (B) ile beklenen (C) bağımsız iki yoldan hesaplandığı için
eşleşmeleri, hesabın doğruluğunun güçlü bir kanıtıdır.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# Motor sabitleri (engine.ts / factors.ts ile birebir)
# ----------------------------------------------------------------------------
GES_KWH_PER_KWP_YIL = 1350
FILO_EV_NET_AZALTIM = 0.70
LED_TASARRUF_ORANI = 0.55
YALITIM_TASARRUF_ORANI = 0.25
KAZAN_TASARRUF_ORANI = 0.12
ORGANIK_ATIK_ORANI = 0.45
KOMPOST_AZALTIM_T_PER_TON = 0.42
AYRISTIRMA_AZALTIM_T_PER_TON = 0.46
GAZ_KWH_PER_M3 = 10.64
KOMUR_KWH_PER_KG = 4.8
GES_CAPEX_TRY_PER_KWP = 28000
GES_SATIS_TARIFE_TRY = 1.8
EV_YAKIT_MALIYET_TASARRUFU = 0.60
CALC_VERSION = "1.1.0"

# Faktör kütüphanesi (factors.ts DEFAULT_FACTORS) — kgCO2e / birim
FAKTORLER = [
    ("ELEKTRIK", 0.442, "kWh", "Kapsam 2"),
    ("DOGALGAZ", 2.02, "m³", "Kapsam 1"),
    ("DIZEL", 2.68, "L", "Kapsam 1"),
    ("JENERATOR_DIZEL", 2.68, "L", "Kapsam 1"),
    ("BENZIN", 2.31, "L", "Kapsam 1"),
    ("LPG", 1.56, "L", "Kapsam 1"),
    ("CNG", 1.90, "m³", "Kapsam 1"),
    ("KOMUR", 1.08, "kg", "Kapsam 1"),
    ("ARAC_KM", 0.192, "km", "Kapsam 1"),
    ("GES_URETIM", 0.442, "kWh", "Kapsam 2 · MAHSUP (kredi)"),
    ("GES_SATIS", 0.442, "kWh", "Bilgi amaçlı (0)"),
    ("ATIK", 580.0, "ton", "Kapsam 3"),
    ("GERI_DONUSUM", 460.0, "ton", "Kapsam 3 · MAHSUP (kredi)"),
    ("KOMPOST", 420.0, "ton", "Kapsam 3 · MAHSUP (kredi)"),
    ("SU", 0.344, "m³", "Kapsam 3"),
    ("UCUS_KM", 0.15, "yolcu-km", "Kapsam 3"),
]

# ----------------------------------------------------------------------------
# Stiller
# ----------------------------------------------------------------------------
FONT_BASLIK = Font(bold=True, size=14, color="1B5E20")
FONT_ALTBASLIK = Font(bold=True, size=11, color="2E7D32")
FONT_BOLD = Font(bold=True)
FONT_KUCUK = Font(italic=True, size=9, color="555555")
FONT_HEADER = Font(bold=True, color="FFFFFF")

FILL_HEADER = PatternFill("solid", fgColor="2E7D32")
FILL_FORMUL = PatternFill("solid", fgColor="FFF9C4")   # açık sarı — canlı formül
FILL_BEKLE = PatternFill("solid", fgColor="C8E6C9")    # açık yeşil — beklenen
FILL_GIRDI = PatternFill("solid", fgColor="E3F2FD")    # açık mavi — girdi
FILL_BAND = PatternFill("solid", fgColor="F1F8E9")

ORTA = Alignment(horizontal="center", vertical="center")
SOL = Alignment(horizontal="left", vertical="center", wrap_text=True)
ded = Side(style="thin", color="BBBBBB")
KENAR = Border(left=ded, right=ded, top=ded, bottom=ded)

NUM4 = "#,##0.0000"
NUM2 = "#,##0.00"
NUM0 = "#,##0"
PCT = "0.00"


def kontrol_formula(r):
    """Sayısal (tolerans 0.01) ve metinsel beklenen değerleri karşılaştıran KONTROL formülü."""
    return (f'=IF(ISNUMBER(C{r}),'
            f'IF(ABS(B{r}-C{r})<=0.01,"✓ EŞLEŞTİ","✗ FARK "&TEXT(B{r}-C{r},"0.000")),'
            f'IF(EXACT(B{r},C{r}),"✓ EŞLEŞTİ","✗ FARK"))')


def sayfa_basligi(ws, baslik, aciklama, kaynak):
    ws.merge_cells("A1:F1")
    c = ws["A1"]; c.value = baslik; c.font = FONT_BASLIK; c.alignment = SOL
    ws.merge_cells("A2:F2")
    c = ws["A2"]; c.value = aciklama; c.font = FONT_KUCUK; c.alignment = SOL
    ws.merge_cells("A3:F3")
    c = ws["A3"]; c.value = f"Kaynak: engine.ts → {kaynak}"; c.font = FONT_KUCUK; c.alignment = SOL
    ws.row_dimensions[2].height = 30
    for col, w in zip("ABCDEF", (34, 18, 16, 22, 48, 26)):
        ws.column_dimensions[col].width = w


def sonuc_baslik(ws, r, kaldiraç="Gösterge"):
    hs = [kaldiraç, "Canlı Sonuç (formül)", "Beklenen", "KONTROL", "Formül (metin)", "Kaynak fonksiyon"]
    for i, h in enumerate(hs, 1):
        c = ws.cell(r, i, h); c.font = FONT_HEADER; c.fill = FILL_HEADER
        c.alignment = ORTA if i != 5 else SOL; c.border = KENAR
    return r + 1


def sonuc_satiri(ws, r, label, formula, expected, source, numfmt=NUM4):
    ws.cell(r, 1, label).alignment = SOL
    b = ws.cell(r, 2, formula); b.fill = FILL_FORMUL; b.number_format = numfmt; b.border = KENAR
    e = ws.cell(r, 3, expected); e.fill = FILL_BEKLE; e.border = KENAR
    if isinstance(expected, (int, float)):
        e.number_format = numfmt
    k = ws.cell(r, 4, kontrol_formula(r)); k.font = FONT_BOLD; k.alignment = ORTA; k.border = KENAR
    disp = formula[1:] if isinstance(formula, str) and formula.startswith("=") else formula
    d = ws.cell(r, 5, disp); d.font = Font(name="Consolas", size=9); d.alignment = SOL; d.border = KENAR
    ws.cell(r, 6, source).font = FONT_KUCUK
    ws.cell(r, 1).border = KENAR
    ws.cell(r, 6).border = KENAR
    return r + 1


def altbaslik(ws, r, metin):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws.cell(r, 1, metin); c.font = FONT_ALTBASLIK; c.fill = FILL_BAND; c.alignment = SOL
    return r + 1


def girdi_satiri(ws, r, label, deger, birim="", numfmt=NUM4):
    ws.cell(r, 1, label).alignment = SOL
    g = ws.cell(r, 2, deger); g.fill = FILL_GIRDI; g.border = KENAR
    if isinstance(deger, (int, float)):
        g.number_format = numfmt
    ws.cell(r, 3, birim).font = FONT_KUCUK
    ws.cell(r, 1).border = KENAR
    return r + 1


wb = Workbook()

# ============================================================================
# 0) KAPAK
# ============================================================================
ws = wb.active
ws.title = "0 · Kapak"
for col, w in zip("ABCDEF", (30, 16, 12, 30, 18, 10)):
    ws.column_dimensions[col].width = w
ws.merge_cells("A1:F1")
c = ws["A1"]; c.value = "Kleaf Karbon Hesaplama — Doğrulama Çalışma Kitabı"
c.font = Font(bold=True, size=16, color="1B5E20")
ws.merge_cells("A2:F2")
c = ws["A2"]; c.value = (f"Hesap sürümü (CALC_VERSION): {CALC_VERSION}  ·  Tüm formüller engine.ts ile birebir. "
                         "Sarı hücreler = CANLI Excel formülü, yeşil = motor beklenen değeri, "
                         "KONTROL sütunu ikisini otomatik karşılaştırır.")
c.font = FONT_KUCUK; c.alignment = SOL
ws.row_dimensions[2].height = 44

r = 4
r = altbaslik(ws, r, "İçindekiler — Hesap Modülleri")
modul_listesi = [
    ("1 · Temel Emisyon", "computeKgCO2e, kgToTons — miktar × faktör, kredi/bilgi kuralları"),
    ("2 · Kapsam + GES Mahsubu", "monthlyScopeTotals — Kapsam 2 aylık MAX(0, ...) kırpması"),
    ("3 · Göstergeler", "yoyChangePct, intensityPerDenominator, targetGapPct"),
    ("4 · Patika + Projeksiyon", "linearNetZeroPath, trendProjection (en küçük kareler)"),
    ("5 · Senaryo Azaltım (t)", "scenarioAnnualSavings + clampPair üst sınır kuralı"),
    ("6 · Senaryo Tasarruf (₺)", "scenarioAnnualSavingsTRY — kWh/m³/L → ₺"),
    ("7 · Yıllıklandırma + Patika", "annualize, scenarioPath (rampalı azaltım)"),
    ("8 · Finansal", "paybackYears, priorityScores (0–100 normalize)"),
    ("9 · GES Fizibilite", "gesFeasibility, gesCoverageRatio"),
    ("10 · Enerji Eşdeğeri", "kwhEquivalent, savingsTargetProgress"),
    ("11 · Veri Kalitesi", "detectAnomalies (medyan+MAD), qualityScore"),
    ("12 · Filo Önceliklendirme", "fleetPriorityScore, yakıt katsayıları"),
]
for ad, aciklama in modul_listesi:
    ws.cell(r, 1, ad).font = FONT_BOLD
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    ws.cell(r, 2, aciklama).alignment = SOL
    r += 1

r += 1
r = altbaslik(ws, r, "Faktör Kütüphanesi (factors.ts · DEFAULT_FACTORS)")
for i, h in enumerate(["Kategori", "Faktör", "Birim", "Kapsam / Not"], 1):
    cc = ws.cell(r, i, h); cc.font = FONT_HEADER; cc.fill = FILL_HEADER; cc.border = KENAR
    if i == 4:
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
r += 1
for ad, fk, birim, kapsam in FAKTORLER:
    ws.cell(r, 1, ad).border = KENAR
    cc = ws.cell(r, 2, fk); cc.number_format = NUM4; cc.border = KENAR
    ws.cell(r, 3, birim).border = KENAR
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
    ws.cell(r, 4, kapsam).alignment = SOL
    r += 1

r += 1
r = altbaslik(ws, r, "Motor Sabitleri (engine.ts)")
sabitler = [
    ("GES_KWH_PER_KWP_YIL", GES_KWH_PER_KWP_YIL, "kWh/kWp·yıl — GES üretim varsayımı"),
    ("FILO_EV_NET_AZALTIM", FILO_EV_NET_AZALTIM, "EV dönüşümü net azaltım oranı"),
    ("LED_TASARRUF_ORANI", LED_TASARRUF_ORANI, "LED aydınlatma tasarruf oranı"),
    ("YALITIM_TASARRUF_ORANI", YALITIM_TASARRUF_ORANI, "Yalıtım tasarruf oranı"),
    ("KAZAN_TASARRUF_ORANI", KAZAN_TASARRUF_ORANI, "Kazan iyileştirme tasarruf oranı"),
    ("ORGANIK_ATIK_ORANI", ORGANIK_ATIK_ORANI, "Kompostlanabilir organik oran"),
    ("KOMPOST_AZALTIM_T_PER_TON", KOMPOST_AZALTIM_T_PER_TON, "tCO2e / ton kompost"),
    ("AYRISTIRMA_AZALTIM_T_PER_TON", AYRISTIRMA_AZALTIM_T_PER_TON, "tCO2e / ton geri dönüşüm"),
    ("GAZ_KWH_PER_M3", GAZ_KWH_PER_M3, "Doğalgaz kWh/m³ dönüşümü"),
    ("KOMUR_KWH_PER_KG", KOMUR_KWH_PER_KG, "Kömür kWh/kg dönüşümü"),
    ("GES_CAPEX_TRY_PER_KWP", GES_CAPEX_TRY_PER_KWP, "GES yatırım ₺/kWp"),
    ("GES_SATIS_TARIFE_TRY", GES_SATIS_TARIFE_TRY, "GES fazla satış tarifesi ₺/kWh"),
    ("EV_YAKIT_MALIYET_TASARRUFU", EV_YAKIT_MALIYET_TASARRUFU, "EV yakıt maliyet tasarruf oranı"),
]
for ad, dg, aciklama in sabitler:
    ws.cell(r, 1, ad).font = FONT_BOLD
    cc = ws.cell(r, 2, dg); cc.number_format = NUM4; cc.fill = FILL_GIRDI
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    ws.cell(r, 3, aciklama).alignment = SOL
    r += 1

# ============================================================================
# 1) TEMEL EMİSYON
# ============================================================================
ws = wb.create_sheet("1 · Temel Emisyon")
sayfa_basligi(
    ws, "1 · Temel Emisyon Hesabı",
    "kg = işaret × miktar × faktör.  İşaret: normal=+1, mahsup(kredi)=−1, bilgi amaçlı=0.  "
    "tCO2e = kg / 1000.  Kural: miktar<0 veya faktör<0 ise motor hata fırlatır.",
    "computeKgCO2e, kgToTons")
r = 5
hdr = ["Kayıt (kategori)", "Miktar", "Faktör", "İşaret", "kg CO2e (canlı)",
       "tCO2e (canlı)", "Beklenen tCO2e", "KONTROL"]
for i, h in enumerate(hdr, 1):
    cc = ws.cell(r, i, h); cc.font = FONT_HEADER; cc.fill = FILL_HEADER; cc.alignment = ORTA; cc.border = KENAR
ws.column_dimensions["A"].width = 30
for col in "BCDEFGH":
    ws.column_dimensions[col].width = 15
kayitlar = [
    ("ELEKTRIK  1.000.000 kWh", 1_000_000, 0.442, 1, 442.0),
    ("DOGALGAZ  10.000 m³", 10_000, 2.02, 1, 20.2),
    ("DIZEL  5.000 L", 5_000, 2.68, 1, 13.4),
    ("BENZIN  2.000 L", 2_000, 2.31, 1, 4.62),
    ("KOMUR  3.000 kg", 3_000, 1.08, 1, 3.24),
    ("CNG  1.000 m³", 1_000, 1.90, 1, 1.90),
    ("SU  0 m³ (sıfır)", 0, 0.344, 1, 0.0),
    ("GES_URETIM  1.000 kWh (mahsup)", 1_000, 0.442, -1, -0.442),
    ("GERI_DONUSUM  10 ton (mahsup)", 10, 460.0, -1, -4.6),
    ("KOMPOST  10 ton (mahsup)", 10, 420.0, -1, -4.2),
    ("GES_SATIS  50.000 kWh (bilgi=0)", 50_000, 0.442, 0, 0.0),
]
r += 1
for ad, miktar, faktor, isaret, beklenen in kayitlar:
    ws.cell(r, 1, ad).border = KENAR
    ws.cell(r, 2, miktar).border = KENAR; ws.cell(r, 2).number_format = NUM0
    ws.cell(r, 3, faktor).border = KENAR; ws.cell(r, 3).number_format = NUM4
    ws.cell(r, 4, isaret).border = KENAR; ws.cell(r, 4).alignment = ORTA
    kg = ws.cell(r, 5, f"=D{r}*B{r}*C{r}"); kg.fill = FILL_FORMUL; kg.number_format = NUM2; kg.border = KENAR
    t = ws.cell(r, 6, f"=E{r}/1000"); t.fill = FILL_FORMUL; t.number_format = NUM4; t.border = KENAR
    e = ws.cell(r, 7, beklenen); e.fill = FILL_BEKLE; e.number_format = NUM4; e.border = KENAR
    k = ws.cell(r, 8, f'=IF(ABS(F{r}-G{r})<=0.001,"✓ EŞLEŞTİ","✗ FARK")')
    k.font = FONT_BOLD; k.alignment = ORTA; k.border = KENAR
    r += 1
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
ws.cell(r, 1, "İşaret sütunu: normal kategoriler +1, kredi kategorileri (GES_URETIM, GERI_DONUSUM, KOMPOST) −1, "
              "bilgi amaçlı (GES_SATIS) 0. Böylece mahsuplar toplam emisyondan düşülür.").font = FONT_KUCUK

# ============================================================================
# 2) KAPSAM + GES MAHSUBU
# ============================================================================
ws = wb.create_sheet("2 · Kapsam+GES")
sayfa_basligi(
    ws, "2 · Kapsam Toplamları ve GES Mahsubu",
    "Her AY için: Kapsam2 = MAX(0, Σscope2 + GES kredileri).  GES mahsubu negatife düşse bile ay bazında 0'da kırpılır "
    "ve BİR SONRAKİ AYA DEVRETMEZ. Toplam = K1 + K2(kırpılmış) + K3.",
    "monthlyScopeTotals")
r = 5
r = altbaslik(ws, r, "Örnek A — Tek ayda aşırı mahsup (Kapsam 2 sıfırda kırpılır)")
hdr = ["Ay", "K1 (ham)", "K2 ham (GES dahil)", "K3 (ham)", "K2 kırpılmış (canlı)",
       "Toplam (canlı)", "Beklenen toplam", "KONTROL"]
for i, h in enumerate(hdr, 1):
    cc = ws.cell(r, i, h); cc.font = FONT_HEADER; cc.fill = FILL_HEADER; cc.alignment = ORTA; cc.border = KENAR
for col in "ABCDEFGH":
    ws.column_dimensions[col].width = 17
ws.column_dimensions["A"].width = 12
r += 1
# Haziran: scope2 100 + GES -130 => -30, s1 40 => K2=0, toplam=40
satirlar_a = [
    ("Haziran", 40, 100 - 130, 0, 40.0),   # -30 -> 0
    ("Mart", 0, 70 - 0, 0, 70.0),          # kısmi mahsup: 100-30=70
]
for ay, k1, k2ham, k3, bekl in satirlar_a:
    ws.cell(r, 1, ay).border = KENAR
    ws.cell(r, 2, k1).border = KENAR; ws.cell(r, 2).number_format = NUM2
    ws.cell(r, 3, k2ham).border = KENAR; ws.cell(r, 3).number_format = NUM2
    ws.cell(r, 4, k3).border = KENAR; ws.cell(r, 4).number_format = NUM2
    kk = ws.cell(r, 5, f"=MAX(0,C{r})"); kk.fill = FILL_FORMUL; kk.number_format = NUM2; kk.border = KENAR
    tt = ws.cell(r, 6, f"=B{r}+E{r}+D{r}"); tt.fill = FILL_FORMUL; tt.number_format = NUM2; tt.border = KENAR
    ee = ws.cell(r, 7, bekl); ee.fill = FILL_BEKLE; ee.number_format = NUM2; ee.border = KENAR
    ws.cell(r, 8, f'=IF(ABS(F{r}-G{r})<=0.001,"✓ EŞLEŞTİ","✗ FARK")').font = FONT_BOLD
    ws.cell(r, 8).alignment = ORTA; ws.cell(r, 8).border = KENAR
    r += 1
r += 1
r = altbaslik(ws, r, "Örnek B — Aylık bağımsızlık (mahsup sonraki aya DEVRETMEZ)")
ws.cell(r, 1, "Ocak K2 ham").border = KENAR
ws.cell(r, 2, 10 - 50).number_format = NUM2; ws.cell(r, 2).border = KENAR
oc = r
ws.cell(r, 3, "Ocak K2 kırpılmış (canlı)").alignment = SOL
kc = ws.cell(r, 4, f"=MAX(0,B{r})"); kc.fill = FILL_FORMUL; kc.number_format = NUM2; kc.border = KENAR
r += 1
ws.cell(r, 1, "Şubat K2 ham").border = KENAR
ws.cell(r, 2, 100).number_format = NUM2; ws.cell(r, 2).border = KENAR
sb = r
ws.cell(r, 3, "Şubat K2 kırpılmış (canlı)").alignment = SOL
ks = ws.cell(r, 4, f"=MAX(0,B{r})"); ks.fill = FILL_FORMUL; ks.number_format = NUM2; ks.border = KENAR
r += 1
ws.cell(r, 1, "YIL K2 = Σ aylık kırpılmış").font = FONT_BOLD
yl = ws.cell(r, 2, f"=D{oc}+D{sb}"); yl.fill = FILL_FORMUL; yl.number_format = NUM2; yl.border = KENAR
ws.cell(r, 3, "Beklenen").font = FONT_BOLD
be = ws.cell(r, 4, 100.0); be.fill = FILL_BEKLE; be.number_format = NUM2; be.border = KENAR
ws.cell(r, 5, f'=IF(ABS(B{r}-D{r})<=0.001,"✓ EŞLEŞTİ","✗ FARK")').font = FONT_BOLD
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
ws.cell(r, 1, "Ocak −40 → 0 (kaybolur), Şubat +100 → 100. Yıl = 0+100 = 100. "
              "Eğer mahsup devretseydi 60 çıkardı; kırpma ay bazında yapıldığı için 100 doğru sonuçtur.").font = FONT_KUCUK

# ============================================================================
# 3) GÖSTERGELER
# ============================================================================
ws = wb.create_sheet("3 · Göstergeler")
sayfa_basligi(
    ws, "3 · Karşılaştırmalı Göstergeler",
    "YoY = (cari−önceki)/önceki×100 (önceki≤0 → null).  Yoğunluk = toplam/payda (payda≤0 → null).  "
    "Hedef sapması = (gerçek−hedef)/hedef×100 (hedef≤0 → null).",
    "yoyChangePct, intensityPerDenominator, targetGapPct")
r = 5
r = altbaslik(ws, r, "Girdiler")
r = girdi_satiri(ws, r, "cari emisyon", 150, "tCO2e"); cari = r - 1
r = girdi_satiri(ws, r, "önceki yıl emisyon", 200, "tCO2e"); onceki = r - 1
r = girdi_satiri(ws, r, "toplam emisyon", 442, "tCO2e"); toplam = r - 1
r = girdi_satiri(ws, r, "payda (üretim vb.)", 400, "birim"); payda = r - 1
r = girdi_satiri(ws, r, "gerçekleşen (hedef için)", 110, "tCO2e"); gercek = r - 1
r = girdi_satiri(ws, r, "hedef", 100, "tCO2e"); hedef = r - 1
r += 1
r = sonuc_baslik(ws, r)
r = sonuc_satiri(ws, r, "YoY değişim %", f"=(B{cari}-B{onceki})/B{onceki}*100", -25.0,
                 "yoyChangePct", PCT)
r = sonuc_satiri(ws, r, "YoY (önceki=0 → null)", '=IF(0<=0,"null",1)', "null", "yoyChangePct")
r = sonuc_satiri(ws, r, "Emisyon yoğunluğu", f"=B{toplam}/B{payda}", 1.105,
                 "intensityPerDenominator", NUM4)
r = sonuc_satiri(ws, r, "Yoğunluk (payda=0 → null)", '=IF(0<=0,"null",1)', "null",
                 "intensityPerDenominator")
r = sonuc_satiri(ws, r, "Hedef sapması % (üstünde)", f"=(B{gercek}-B{hedef})/B{hedef}*100", 10.0,
                 "targetGapPct", PCT)
r = sonuc_satiri(ws, r, "Hedef sapması % (altında)", f"=(90-B{hedef})/B{hedef}*100", -10.0,
                 "targetGapPct", PCT)

# ============================================================================
# 4) PATİKA + PROJEKSİYON
# ============================================================================
ws = wb.create_sheet("4 · Patika+Projeksiyon")
sayfa_basligi(
    ws, "4 · Net-Sıfır Patikası ve Trend Projeksiyonu",
    "Doğrusal patika: hedef(y) = baz × (1 − (y−bazYıl)/(netZeroYıl−bazYıl)).  "
    "Trend: en küçük kareler; eğim=(nΣxy−ΣxΣy)/(nΣxx−Σx²), kesişim=(Σy−eğim·Σx)/n, tahmin=MAX(0, eğim·y+kesişim).",
    "linearNetZeroPath, trendProjection")
r = 5
r = altbaslik(ws, r, "Doğrusal Net-Sıfır Patikası")
r = girdi_satiri(ws, r, "baz yıl", 2024, "", NUM0); bazYil = r - 1
r = girdi_satiri(ws, r, "baz emisyon", 1000, "tCO2e", NUM0); bazEm = r - 1
r = girdi_satiri(ws, r, "net-sıfır yılı", 2053, "", NUM0); nzYil = r - 1
r += 1
r = sonuc_baslik(ws, r, "Yıl")
for yil, bekl in [(2024, 1000.0), (2025, 965.5172), (2026, 931.0345),
                  (2038, 517.2414), (2053, 0.0)]:
    f = f"=B{bazEm}*(1-({yil}-B{bazYil})/(B{nzYil}-B{bazYil}))"
    r = sonuc_satiri(ws, r, f"hedef({yil})", f, bekl, "linearNetZeroPath", NUM4)
r += 1
r = altbaslik(ws, r, "Trend Projeksiyonu — geçmiş: (2023,300) (2024,250) (2025,200)")
# x = yıl, y değerleri
x1, y1r = r, r
for yil_x, y_val in [(2023, 300), (2024, 250), (2025, 200)]:
    ax = ws.cell(r, 1, yil_x); ax.border = KENAR; ax.number_format = NUM0
    bx = ws.cell(r, 2, y_val); bx.fill = FILL_GIRDI; bx.border = KENAR
    r += 1
x2 = r - 1
ws.cell(x1, 3, "← geçmiş y değerleri (B sütunu), x = yıl (A sütunu)").font = FONT_KUCUK
r += 1
# SLOPE/INTERCEPT ile çapraz kontrol
r = sonuc_baslik(ws, r, "Trend")
slope_ref = f"SLOPE(B{x1}:B{x2},A{x1}:A{x2})"
inter_ref = f"INTERCEPT(B{x1}:B{x2},A{x1}:A{x2})"
r = sonuc_satiri(ws, r, "Eğim (slope)", f"={slope_ref}", -50.0, "en küçük kareler", NUM2)
r = sonuc_satiri(ws, r, "Kesişim (intercept)", f"={inter_ref}", 101450.0, "en küçük kareler", NUM2)
for yil, bekl in [(2026, 150.0), (2027, 100.0), (2030, 0.0), (2032, 0.0)]:
    f = f"=MAX(0,{slope_ref}*{yil}+{inter_ref})"
    r = sonuc_satiri(ws, r, f"tahmin({yil})", f, bekl, "trendProjection", NUM2)
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
ws.cell(r, 1, "Eğim −50/yıl (300→200, 2 yıl). 2030 ve sonrası negatif çıktığından MAX(0,…) ile 0'a kırpılır.").font = FONT_KUCUK

# ============================================================================
# 5) SENARYO AZALTIM (t)
# ============================================================================
ws = wb.create_sheet("5 · Senaryo Azaltım")
sayfa_basligi(
    ws, "5 · Senaryo Yıllık Azaltım (tCO2e)",
    "pct(v)=MIN(100,MAX(0,v))/100.  GES=kWp×1350×fEl/1000; Bina=binaT×pct; Filo=filoT×pct×0.70; "
    "LED=kWh×pct×0.55×fEl/1000; Kompost=atıkT×0.45×pct×0.42; Ayrıştırma=atıkT×0.55×pct×0.46.  "
    "Yalıtım/Kazan ve Filo/Toplu çiftleri clampPair ile üst sınıra kırpılır.",
    "scenarioAnnualSavings, clampPair")
r = 5
r = altbaslik(ws, r, "Bağlam (context) girdileri")
r = girdi_satiri(ws, r, "fEl — elektrik faktörü", 0.442, "kg/kWh"); fEl = r - 1
r = girdi_satiri(ws, r, "filoTCO2e", 200, "tCO2e", NUM0); filoT = r - 1
r = girdi_satiri(ws, r, "binaEnerjiTCO2e", 600, "tCO2e", NUM0); binaT = r - 1
r = girdi_satiri(ws, r, "aydinlatmaKwh", 1_000_000, "kWh", NUM0); aydK = r - 1
r = girdi_satiri(ws, r, "dogalgazTCO2e", 300, "tCO2e", NUM0); gazT = r - 1
r = girdi_satiri(ws, r, "atikTon", 1000, "ton", NUM0); atikT = r - 1
r = girdi_satiri(ws, r, "kWh/kWp·yıl sabiti", 1350, "", NUM0); kwpSbt = r - 1
r += 1
r = altbaslik(ws, r, "Kaldıraçlar (izole)")
r = sonuc_baslik(ws, r, "Kaldıraç / girdi")
# GES 1000 kWp
r = sonuc_satiri(ws, r, "GES — 1000 kWp", f"=1000*B{kwpSbt}*B{fEl}/1000", 596.7,
                 "scenarioAnnualSavings", NUM4)
r = sonuc_satiri(ws, r, "Bina — %10", f"=B{binaT}*MIN(100,MAX(0,10))/100", 60.0,
                 "scenarioAnnualSavings", NUM4)
r = sonuc_satiri(ws, r, "Filo — %50 (×0.70)", f"=B{filoT}*MIN(100,MAX(0,50))/100*0.7", 70.0,
                 "scenarioAnnualSavings", NUM4)
r = sonuc_satiri(ws, r, "LED — %100 (×0.55)", f"=B{aydK}*MIN(100,MAX(0,100))/100*0.55*B{fEl}/1000",
                 243.1, "scenarioAnnualSavings", NUM4)
r = sonuc_satiri(ws, r, "Kompost — %50", f"=B{atikT}*0.45*MIN(100,MAX(0,50))/100*0.42", 94.5,
                 "scenarioAnnualSavings", NUM4)
r = sonuc_satiri(ws, r, "Ayrıştırma — %50", f"=B{atikT}*0.55*MIN(100,MAX(0,50))/100*0.46", 126.5,
                 "scenarioAnnualSavings", NUM4)
def clamp_blok(ws, r, baslik, hdr_ilk, raw_f1, raw_f2, cap_f, bek1, bek2, numfmt, aciklama):
    """Tek satırlık clampPair gösterimi: ham1|ham2|toplam|üstsınır|k|son1|son2|bekl1|bekl2|KONTROL."""
    r = altbaslik(ws, r, baslik)
    hdr = ["", hdr_ilk[0], hdr_ilk[1], "Toplam", "Üst sınır", "k çarpanı",
           hdr_ilk[2], hdr_ilk[3], "Bekl. " + hdr_ilk[2], "Bekl. " + hdr_ilk[3], "KONTROL"]
    for i, h in enumerate(hdr, 1):
        cc = ws.cell(r, i, h); cc.font = FONT_HEADER; cc.fill = FILL_HEADER; cc.alignment = ORTA; cc.border = KENAR
    r += 1
    d = r
    ws.cell(d, 1, "%100 / %100").font = FONT_BOLD
    c2 = ws.cell(d, 2, raw_f1); c2.fill = FILL_FORMUL; c2.number_format = numfmt
    c3 = ws.cell(d, 3, raw_f2); c3.fill = FILL_FORMUL; c3.number_format = numfmt
    ws.cell(d, 4, f"=B{d}+C{d}").number_format = numfmt
    ws.cell(d, 5, cap_f).number_format = numfmt
    ws.cell(d, 6, f"=IF(OR(D{d}<=E{d},D{d}<=0),1,E{d}/D{d})").number_format = NUM4
    c7 = ws.cell(d, 7, f"=B{d}*F{d}"); c7.fill = FILL_FORMUL; c7.number_format = numfmt
    c8 = ws.cell(d, 8, f"=C{d}*F{d}"); c8.fill = FILL_FORMUL; c8.number_format = numfmt
    e9 = ws.cell(d, 9, bek1); e9.fill = FILL_BEKLE; e9.number_format = numfmt
    e10 = ws.cell(d, 10, bek2); e10.fill = FILL_BEKLE; e10.number_format = numfmt
    kk = ws.cell(d, 11, f'=IF(AND(ABS(G{d}-I{d})<=0.01,ABS(H{d}-J{d})<=0.01),"✓ EŞLEŞTİ","✗ FARK")')
    kk.font = FONT_BOLD; kk.alignment = ORTA
    for i in range(1, 12):
        ws.cell(d, i).border = KENAR
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
    ws.cell(r, 1, aciklama).font = FONT_KUCUK
    return r + 2


r += 1
r = clamp_blok(
    ws, r, "clampPair — Yalıtım + Kazan (üst sınır = dogalgazTCO2e = 300)",
    ("Yalıtım ham", "Kazan ham", "Yalıtım son", "Kazan son"),
    f"=B{gazT}*1*0.25", f"=B{gazT}*1*0.12", f"=B{gazT}", 75.0, 36.0, NUM2,
    "Yalıtım=300×0.25=75, Kazan=300×0.12=36. Toplam 111 ≤ 300 üst sınır → kırpma yok (k=1).")
fs_bek = round(200 * 140 / 340, 4)
ts_bek = round(200 * 200 / 340, 4)
r = clamp_blok(
    ws, r, "clampPair — Filo(%100)+Toplu(%100) (üst sınır = filoTCO2e = 200 → KIRPILIR)",
    ("Filo ham", "Toplu ham", "Filo son", "Toplu son"),
    f"=B{filoT}*1*0.7", f"=B{filoT}*1", f"=B{filoT}", fs_bek, ts_bek, NUM4,
    "Ham toplam 140+200=340 > 200 → k=200/340=0.5882; Filo=82.35, Toplu=117.65, toplam tam 200.")
for col in "BCDEFGHIJK":
    ws.column_dimensions[col].width = 13

# ============================================================================
# 6) SENARYO TASARRUF (₺)
# ============================================================================
ws = wb.create_sheet("6 · Senaryo TL")
sayfa_basligi(
    ws, "6 · Senaryo Yıllık Tasarruf (₺)",
    "GES₺=kWp×1350×elFiyat; LED₺=kWh×pct×0.55×elFiyat; Gaz₺=(yalıtım+kazan)×1000/gazFaktör×gazFiyat; "
    "Filo₺=filoL×dizelFiyat×0.60+topluL×dizelFiyat; Atık₺=saptırılanTon×bertarafFiyat.",
    "scenarioAnnualSavingsTRY")
r = 5
r = altbaslik(ws, r, "Girdiler ve fiyatlar")
r = girdi_satiri(ws, r, "GES kWp", 100, "kWp", NUM0); gKwp = r - 1
r = girdi_satiri(ws, r, "aydinlatmaKwh", 1_000_000, "kWh", NUM0); aK = r - 1
r = girdi_satiri(ws, r, "LED %", 10, "%", NUM0); ledP = r - 1
r = girdi_satiri(ws, r, "elektrik fiyatı", 4, "₺/kWh", NUM2); elF = r - 1
r = girdi_satiri(ws, r, "kWh/kWp·yıl sabiti", 1350, "", NUM0); kSbt = r - 1
r += 1
r = sonuc_baslik(ws, r)
r = sonuc_satiri(ws, r, "GES tasarrufu ₺", f"=B{gKwp}*B{kSbt}*B{elF}", 540000.0,
                 "scenarioAnnualSavingsTRY", NUM0)
r = sonuc_satiri(ws, r, "LED tasarrufu ₺",
                 f"=B{aK}*MIN(100,MAX(0,B{ledP}))/100*0.55*B{elF}", 220000.0,
                 "scenarioAnnualSavingsTRY", NUM0)
r += 1
r = altbaslik(ws, r, "Gaz / Filo / Atık ₺ — faktör dönüşümleri")
r = girdi_satiri(ws, r, "yalıtım tCO2e (örnek)", 75, "tCO2e", NUM0); yTon = r - 1
r = girdi_satiri(ws, r, "kazan tCO2e (örnek)", 36, "tCO2e", NUM0); kTon = r - 1
r = girdi_satiri(ws, r, "gaz faktörü", 2.02, "kg/m³", NUM4); gazFk = r - 1
r = girdi_satiri(ws, r, "gaz fiyatı", 17, "₺/m³", NUM2); gazFy = r - 1
r = girdi_satiri(ws, r, "filo tCO2e (azaltım)", 70, "tCO2e", NUM0); filoTon = r - 1
r = girdi_satiri(ws, r, "toplu taşıma tCO2e", 40, "tCO2e", NUM0); topluTon = r - 1
r = girdi_satiri(ws, r, "dizel faktörü", 2.68, "kg/L", NUM4); dzFk = r - 1
r = girdi_satiri(ws, r, "dizel fiyatı", 52, "₺/L", NUM2); dzFy = r - 1
r = girdi_satiri(ws, r, "saptırılan atık", 500, "ton", NUM0); sapT = r - 1
r = girdi_satiri(ws, r, "bertaraf fiyatı", 850, "₺/ton", NUM0); berF = r - 1
r += 1
r = sonuc_baslik(ws, r)
gazTRY = (75 + 36) * 1000 / 2.02 * 17
r = sonuc_satiri(ws, r, "Gaz tasarrufu ₺",
                 f"=(B{yTon}+B{kTon})*1000/B{gazFk}*B{gazFy}", round(gazTRY, 2),
                 "scenarioAnnualSavingsTRY", NUM0)
filoL = (70 / 0.7) * 1000 / 2.68
topluL = 40 * 1000 / 2.68
filoTRY = filoL * 52 * 0.6 + topluL * 52
r = sonuc_satiri(ws, r, "Filo+Toplu tasarrufu ₺",
                 f"=(B{filoTon}/0.7)*1000/B{dzFk}*B{dzFy}*0.6+B{topluTon}*1000/B{dzFk}*B{dzFy}",
                 round(filoTRY, 2), "scenarioAnnualSavingsTRY", NUM0)
r = sonuc_satiri(ws, r, "Atık tasarrufu ₺", f"=B{sapT}*B{berF}", 425000.0,
                 "scenarioAnnualSavingsTRY", NUM0)

# ============================================================================
# 7) YILLIKLANDIRMA + PATİKA
# ============================================================================
ws = wb.create_sheet("7 · Yıllıklandırma")
sayfa_basligi(
    ws, "7 · Yıllıklandırma ve Senaryo Patikası",
    "annualize(v,n): n≤0→0; n≥12→v; aksi halde v×12/n.  "
    "scenarioPath: değer − azaltım×MIN(1,(y−başlangıç+1)/rampa), 0'ın altına düşmez.",
    "annualize, scenarioPath")
r = 5
r = altbaslik(ws, r, "annualize(v, n)")
r = sonuc_baslik(ws, r)
for v, n, bekl in [(100, 6, 200.0), (340, 12, 340.0), (340, 14, 340.0),
                   (500, 0, 0.0), (500, -3, 0.0), (10, 1, 120.0), (100, 7, 100 * 12 / 7)]:
    f = f"=IF({n}<=0,0,IF({n}>=12,{v},{v}*12/{n}))"
    r = sonuc_satiri(ws, r, f"annualize({v}, {n})", f, round(bekl, 4), "annualize", NUM4)
r += 1
r = altbaslik(ws, r, "scenarioPath — baz=100/yıl, başlangıç=2026, yıllık azaltım=50, rampa=5")
r = girdi_satiri(ws, r, "baz değer", 100, "tCO2e", NUM0); pv = r - 1
r = girdi_satiri(ws, r, "yıllık azaltım", 50, "tCO2e", NUM0); pa = r - 1
r = girdi_satiri(ws, r, "başlangıç yılı", 2026, "", NUM0); pf = r - 1
r = girdi_satiri(ws, r, "rampa (yıl)", 5, "", NUM0); pr = r - 1
r += 1
r = sonuc_baslik(ws, r, "Yıl")
for yil, bekl in [(2026, 90.0), (2027, 80.0), (2028, 70.0), (2029, 60.0),
                  (2030, 50.0), (2031, 50.0)]:
    f = f"=MAX(0,B{pv}-B{pa}*MIN(1,({yil}-B{pf}+1)/B{pr}))"
    r = sonuc_satiri(ws, r, f"patika({yil})", f, bekl, "scenarioPath", NUM4)
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
ws.cell(r, 1, "Rampa: (yıl−2026+1)/5 → 0.2, 0.4, 0.6, 0.8, 1.0, 1.0. Azaltım tam rampada 50'ye ulaşır ve sabit kalır.").font = FONT_KUCUK

# ============================================================================
# 8) FİNANSAL
# ============================================================================
ws = wb.create_sheet("8 · Finansal")
sayfa_basligi(
    ws, "8 · Geri Ödeme ve Öncelik Skoru",
    "paybackYears: tasarruf≤0→null; capex≤0→0; aksi halde capex/tasarruf.  "
    "priorityScores: oran=azaltım/capex (azaltım≤0→0, capex≤0→∞); skor = ∞→100, "
    "aksi halde MIN(100, oran/enBüyükSonlu×100).",
    "paybackYears, priorityScores")
r = 5
r = altbaslik(ws, r, "paybackYears")
r = sonuc_baslik(ws, r)
r = sonuc_satiri(ws, r, "capex 1.000.000 / tasarruf 250.000",
                 "=IF(250000<=0,\"null\",IF(1000000<=0,0,1000000/250000))", 4.0,
                 "paybackYears", NUM4)
r = sonuc_satiri(ws, r, "tasarruf 0 → null",
                 "=IF(0<=0,\"null\",1)", "null", "paybackYears")
r = sonuc_satiri(ws, r, "capex 0 → 0",
                 "=IF(100<=0,\"null\",IF(0<=0,0,1))", 0.0, "paybackYears", NUM4)
r += 1
r = altbaslik(ws, r, "priorityScores — 4 önlem (azaltım tCO2e, capex ₺)")
hdr = ["Önlem", "Azaltım", "Capex", "Oran (canlı)", "Skor (canlı)", "Beklenen skor", "KONTROL"]
for i, h in enumerate(hdr, 1):
    cc = ws.cell(r, i, h); cc.font = FONT_HEADER; cc.fill = FILL_HEADER; cc.alignment = ORTA; cc.border = KENAR
for col in "ABCDEFG":
    ws.column_dimensions[col].width = 15
ws.column_dimensions["A"].width = 22
r += 1
onlemler = [
    ("A", 100, 1_000_000, 100.0),
    ("B", 50, 1_000_000, 50.0),
    ("C (bedava, capex=0)", 10, 0, 100.0),  # ∞ → 100
    ("D (azaltım=0)", 0, 500, 0.0),
]
ilk = r
for ad, az, cx, bekl in onlemler:
    ws.cell(r, 1, ad).border = KENAR
    ws.cell(r, 2, az).border = KENAR; ws.cell(r, 2).number_format = NUM0
    ws.cell(r, 3, cx).border = KENAR; ws.cell(r, 3).number_format = NUM0
    # oran: azaltım<=0 → 0; capex<=0 → 1E+99 (∞ sentinel)
    orn = ws.cell(r, 4, f"=IF(B{r}<=0,0,IF(C{r}<=0,1E+99,B{r}/C{r}))")
    orn.fill = FILL_FORMUL; orn.number_format = "0.00E+00"; orn.border = KENAR
    r += 1
son = r - 1
# maxFinite = en büyük sonlu oran (1E+90 altındakiler)
for idx in range(ilk, son + 1):
    skor = ws.cell(idx, 5,
                   f'=IF(D{idx}>=1E+90,100,IF(MAXIFS($D${ilk}:$D${son},$D${ilk}:$D${son},"<1E+90")<=0,0,'
                   f'MIN(100,D{idx}/MAXIFS($D${ilk}:$D${son},$D${ilk}:$D${son},"<1E+90")*100)))')
    skor.fill = FILL_FORMUL; skor.number_format = NUM2; skor.border = KENAR
    bekl = onlemler[idx - ilk][3]
    e = ws.cell(idx, 6, bekl); e.fill = FILL_BEKLE; e.number_format = NUM2; e.border = KENAR
    kk = ws.cell(idx, 7, f'=IF(ABS(E{idx}-F{idx})<=0.01,"✓ EŞLEŞTİ","✗ FARK")')
    kk.font = FONT_BOLD; kk.alignment = ORTA; kk.border = KENAR
r = son + 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
ws.cell(r, 1, "En büyük sonlu oran A önlemi (0.0001). Bedava önlem (capex 0) sonsuz verim → 100. "
              "Azaltımı olmayan önlem → 0.").font = FONT_KUCUK

# ============================================================================
# 9) GES FİZİBİLİTE
# ============================================================================
ws = wb.create_sheet("9 · GES Fizibilite")
sayfa_basligi(
    ws, "9 · GES Fizibilite ve Karşılama Oranı",
    "uretim=kWp×1350; özTüketim=uretim×pct; satış=uretim−öz; azaltım=öz×fEl/1000; "
    "gelir=öz×elFiyat+satış×satışTarife(1.8); capex=verilen>0?verilen:kWp×28000; payback=capex/gelir.  "
    "Karşılama = ges/(ges+şebeke)×100.",
    "gesFeasibility, gesCoverageRatio")
r = 5
r = altbaslik(ws, r, "Girdiler")
r = girdi_satiri(ws, r, "kWp", 500, "kWp", NUM0); kwp = r - 1
r = girdi_satiri(ws, r, "öz tüketim %", 70, "%", NUM0); ozP = r - 1
r = girdi_satiri(ws, r, "fEl", 0.442, "kg/kWh"); fe = r - 1
r = girdi_satiri(ws, r, "elektrik fiyatı", 3.0, "₺/kWh", NUM2); ef = r - 1
r = girdi_satiri(ws, r, "satış tarifesi", 1.8, "₺/kWh", NUM2); st = r - 1
r = girdi_satiri(ws, r, "kWh/kWp sabiti", 1350, "", NUM0); ks2 = r - 1
r = girdi_satiri(ws, r, "capex ₺/kWp sabiti", 28000, "", NUM0); cps = r - 1
r += 1
r = sonuc_baslik(ws, r)
uretim = 500 * 1350
oz = uretim * 0.7
satis = uretim - oz
azaltim = oz * 0.442 / 1000
gelir = oz * 3.0 + satis * 1.8
capex = 500 * 28000
payback = capex / gelir
r = sonuc_satiri(ws, r, "Üretim (kWh/yıl)", f"=B{kwp}*B{ks2}", uretim, "gesFeasibility", NUM0)
uretimRow = r - 1
r = sonuc_satiri(ws, r, "Öz tüketim (kWh)", f"=B{uretimRow}*MIN(100,MAX(0,B{ozP}))/100", oz,
                 "gesFeasibility", NUM0)
ozRow = r - 1
r = sonuc_satiri(ws, r, "Satış (kWh)", f"=B{uretimRow}-B{ozRow}", satis, "gesFeasibility", NUM0)
satisRow = r - 1
r = sonuc_satiri(ws, r, "Azaltım (tCO2e)", f"=B{ozRow}*B{fe}/1000", round(azaltim, 4),
                 "gesFeasibility", NUM4)
r = sonuc_satiri(ws, r, "Gelir (₺/yıl)", f"=B{ozRow}*B{ef}+B{satisRow}*B{st}", gelir,
                 "gesFeasibility", NUM0)
gelirRow = r - 1
r = sonuc_satiri(ws, r, "Capex (₺)", f"=B{kwp}*B{cps}", capex, "gesFeasibility", NUM0)
capexRow = r - 1
r = sonuc_satiri(ws, r, "Geri ödeme (yıl)", f"=B{capexRow}/B{gelirRow}", round(payback, 4),
                 "gesFeasibility", NUM4)
r += 1
r = altbaslik(ws, r, "gesCoverageRatio — karşılama %")
r = girdi_satiri(ws, r, "GES üretimi (kWh)", 675000, "kWh", NUM0); gk = r - 1
r = girdi_satiri(ws, r, "şebeke tüketimi (kWh)", 800000, "kWh", NUM0); sk = r - 1
r += 1
r = sonuc_baslik(ws, r)
cover = 675000 / (675000 + 800000) * 100
r = sonuc_satiri(ws, r, "Karşılama oranı %", f"=B{gk}/(B{gk}+B{sk})*100", round(cover, 4),
                 "gesCoverageRatio", PCT)

# ============================================================================
# 10) ENERJİ EŞDEĞERİ
# ============================================================================
ws = wb.create_sheet("10 · Enerji Eşdeğeri")
sayfa_basligi(
    ws, "10 · kWh Eşdeğeri ve Hedef İlerlemesi",
    "kWhEşdeğeri = elektrik + doğalgaz×10.64 + kömür×4.8.  "
    "Hedef ilerlemesi = ((baz−cari)/baz) / (hedef%/100) × 100 (baz≤0 veya hedef≤0 → null).",
    "kwhEquivalent, savingsTargetProgress")
r = 5
r = altbaslik(ws, r, "kWh Eşdeğeri")
r = girdi_satiri(ws, r, "elektrik (kWh)", 120000, "kWh", NUM0); el = r - 1
r = girdi_satiri(ws, r, "doğalgaz (m³)", 5000, "m³", NUM0); dg = r - 1
r = girdi_satiri(ws, r, "kömür (kg)", 2000, "kg", NUM0); km = r - 1
r += 1
r = sonuc_baslik(ws, r)
r = sonuc_satiri(ws, r, "kWh eşdeğeri", f"=B{el}+B{dg}*10.64+B{km}*4.8", 182800.0,
                 "kwhEquivalent", NUM2)
r += 1
r = altbaslik(ws, r, "Tasarruf Hedefi İlerlemesi")
r = girdi_satiri(ws, r, "baz emisyon", 200000, "tCO2e", NUM0); bz = r - 1
r = girdi_satiri(ws, r, "cari emisyon", 182800, "tCO2e", NUM0); cr = r - 1
r = girdi_satiri(ws, r, "hedef azaltım %", 10, "%", NUM0); hd = r - 1
r += 1
r = sonuc_baslik(ws, r)
prog = ((200000 - 182800) / 200000) / (10 / 100) * 100
r = sonuc_satiri(ws, r, "Hedef ilerlemesi %",
                 f"=((B{bz}-B{cr})/B{bz})/(B{hd}/100)*100", round(prog, 4),
                 "savingsTargetProgress", PCT)
r = sonuc_satiri(ws, r, "İlerleme (baz=0 → null)", '=IF(0<=0,"null",1)', "null",
                 "savingsTargetProgress")

# ============================================================================
# 11) VERİ KALİTESİ
# ============================================================================
ws = wb.create_sheet("11 · Veri Kalitesi")
sayfa_basligi(
    ws, "11 · Anomali Tespiti ve Kalite Skoru",
    "detectAnomalies: m=medyan; MAD=medyan(|x−m|); eşik=MAX(3×1.4826×MAD, 0.3×m); "
    "|x−m|>eşik ise anomali; >2×eşik ise 'yuksek'.  "
    "qualityScore = ROUND(100×(0.4×tamlık+0.3×onay+0.2×belge+0.1×(1−aykırı))).",
    "detectAnomalies, qualityScore")
r = 5
r = altbaslik(ws, r, "Anomali serisi: 100, 105, 98, 102, 300, 101")
seri = [100, 105, 98, 102, 300, 101]
hdr = ["Değer", "|x − medyan| (canlı)", "Sapma % (canlı)", "Anomali? (canlı)"]
for i, h in enumerate(hdr, 1):
    cc = ws.cell(r, i, h); cc.font = FONT_HEADER; cc.fill = FILL_HEADER; cc.alignment = ORTA; cc.border = KENAR
ws.column_dimensions["A"].width = 14
for col in "BCD":
    ws.column_dimensions[col].width = 20
r += 1
seri_ilk = r
for v in seri:
    ws.cell(r, 1, v).border = KENAR; ws.cell(r, 1).fill = FILL_GIRDI
    r += 1
seri_son = r - 1
med_ref = f"MEDIAN($A${seri_ilk}:$A${seri_son})"
# |x-med| yardımcı sütun
for idx in range(seri_ilk, seri_son + 1):
    ab = ws.cell(idx, 2, f"=ABS(A{idx}-{med_ref})"); ab.fill = FILL_FORMUL; ab.number_format = NUM4; ab.border = KENAR
mad_ref = f"MEDIAN($B${seri_ilk}:$B${seri_son})"
esik_ref = f"MAX(3*1.4826*{mad_ref},0.3*{med_ref})"
for idx in range(seri_ilk, seri_son + 1):
    sp = ws.cell(idx, 3, f"=(A{idx}-{med_ref})/{med_ref}*100"); sp.fill = FILL_FORMUL; sp.number_format = PCT; sp.border = KENAR
    an = ws.cell(idx, 4, f'=IF(B{idx}>{esik_ref},IF(B{idx}>2*{esik_ref},"yuksek","orta"),"—")')
    an.fill = FILL_FORMUL; an.alignment = ORTA; an.border = KENAR
r = seri_son + 2
r = sonuc_baslik(ws, r)
r = sonuc_satiri(ws, r, "Medyan (m)", f"={med_ref}", 101.5, "detectAnomalies", NUM4)
r = sonuc_satiri(ws, r, "MAD", f"={mad_ref}", 2.5, "detectAnomalies", NUM4)
esik_bek = max(3 * 1.4826 * 2.5, 0.3 * 101.5)
r = sonuc_satiri(ws, r, "Eşik", f"={esik_ref}", round(esik_bek, 4), "detectAnomalies", NUM4)
sapma300 = (300 - 101.5) / 101.5 * 100
r = sonuc_satiri(ws, r, "300 için sapma %", f"=(300-{med_ref})/{med_ref}*100", round(sapma300, 4),
                 "detectAnomalies", PCT)
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
ws.cell(r, 1, "Eşik = MAX(3×1.4826×2.5=11.12, 0.3×101.5=30.45) = 30.45. "
              "300 için |x−m|=198.5 > 2×30.45=60.9 → 'yuksek'. Diğerleri eşik altında → '—'.").font = FONT_KUCUK
r += 2
r = altbaslik(ws, r, "qualityScore")
r = girdi_satiri(ws, r, "beklenen kayıt", 126, "adet", NUM0); qb = r - 1
r = girdi_satiri(ws, r, "dolu (girilmiş)", 125, "adet", NUM0); qd = r - 1
r = girdi_satiri(ws, r, "toplam kayıt", 160, "adet", NUM0); qt = r - 1
r = girdi_satiri(ws, r, "onaylı", 160, "adet", NUM0); qo = r - 1
r = girdi_satiri(ws, r, "belgeli", 160, "adet", NUM0); qbg = r - 1
r = girdi_satiri(ws, r, "aykırı (anomali)", 1, "adet", NUM0); qa = r - 1
r += 1
r = sonuc_baslik(ws, r)
f = (f"=ROUND(100*(0.4*MIN(1,B{qd}/B{qb})+0.3*MIN(1,B{qo}/B{qt})+"
     f"0.2*MIN(1,B{qbg}/B{qt})+0.1*(1-MIN(1,B{qa}/B{qt}))),0)")
r = sonuc_satiri(ws, r, "Kalite skoru (0–100)", f, 100.0, "qualityScore", NUM0)

# ============================================================================
# 12) FİLO ÖNCELİKLENDİRME
# ============================================================================
ws = wb.create_sheet("12 · Filo")
sayfa_basligi(
    ws, "12 · Filo Önceliklendirme Skoru",
    "fleetPriorityScore = MAX(0, tCO2e) × yakıtKatsayısı.  "
    "Katsayılar: DIZEL 1.0, BENZIN 0.9, LPG 0.8, CNG 0.7, ELEKTRIK 0.0, diğer 1.0.",
    "fleetPriorityScore, FUEL_PRIORITY_COEF")
r = 5
r = altbaslik(ws, r, "Yakıt katsayı tablosu")
katsayilar = [("DIZEL", 1.0), ("BENZIN", 0.9), ("LPG", 0.8), ("CNG", 0.7), ("ELEKTRIK", 0.0)]
kat_ilk = r
for yk, kt in katsayilar:
    ws.cell(r, 1, yk).border = KENAR
    ws.cell(r, 2, kt).border = KENAR; ws.cell(r, 2).number_format = NUM2; ws.cell(r, 2).fill = FILL_GIRDI
    r += 1
kat_son = r - 1
r += 1
hdr = ["Araç (yakıt)", "tCO2e", "Katsayı (VLOOKUP)", "Skor (canlı)", "Beklenen", "KONTROL"]
for i, h in enumerate(hdr, 1):
    cc = ws.cell(r, i, h); cc.font = FONT_HEADER; cc.fill = FILL_HEADER; cc.alignment = ORTA; cc.border = KENAR
for col in "ABCDEF":
    ws.column_dimensions[col].width = 18
r += 1
araclar = [("DIZEL", 120, 120.0), ("BENZIN", 120, 108.0), ("LPG", 100, 80.0),
           ("CNG", 90, 63.0), ("ELEKTRIK", 50, 0.0), ("ARAC_KM (diğer→1.0)", 40, 40.0)]
for yk, tc, bekl in araclar:
    ws.cell(r, 1, yk).border = KENAR
    ws.cell(r, 2, tc).border = KENAR; ws.cell(r, 2).number_format = NUM0
    kt = ws.cell(r, 3, f'=IFERROR(VLOOKUP(A{r},$A${kat_ilk}:$B${kat_son},2,FALSE),1)')
    kt.number_format = NUM2; kt.border = KENAR
    sk = ws.cell(r, 4, f"=MAX(0,B{r})*C{r}"); sk.fill = FILL_FORMUL; sk.number_format = NUM2; sk.border = KENAR
    e = ws.cell(r, 5, bekl); e.fill = FILL_BEKLE; e.number_format = NUM2; e.border = KENAR
    kk = ws.cell(r, 6, f'=IF(ABS(D{r}-E{r})<=0.01,"✓ EŞLEŞTİ","✗ FARK")')
    kk.font = FONT_BOLD; kk.alignment = ORTA; kk.border = KENAR
    r += 1
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
ws.cell(r, 1, "Katsayı, yakıt adıyla tablodan VLOOKUP (tam eşleşme) ile çekilir; tabloda olmayan yakıt (ör. ARAC_KM) "
              "için IFERROR ile varsayılan 1.0 kullanılır. ELEKTRIK katsayısı 0 → skor 0.").font = FONT_KUCUK

# ----------------------------------------------------------------------------
import os
cikti = os.path.join(os.path.dirname(os.path.abspath(__file__)), "kleaf-hesap-dogrulama.xlsx")
wb.save(cikti)
print("OLUSTURULDU:", cikti)
print("Sayfalar:", " | ".join(wb.sheetnames))
print("Boyut:", os.path.getsize(cikti), "bayt")
